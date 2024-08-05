from openpyxl.reader.excel import load_workbook


def read_programs_list():
    wb = load_workbook('kru_full.xlsx')
    sheet = wb.active
    # sheet = wb.active
    # wb.get_sheet_by_name("Details")
    data = []
    for row in sheet.iter_rows(min_row=4, max_row=2333, min_col=1, max_col=8, values_only=True):
        data.append(row)
    return data


exel_list = read_programs_list()
limit = 10
temp_list = []
mem = str()
for i in range(len(exel_list)):
    if exel_list[i][0] == exel_list[i + 2][0]:
        j = i
        temp_list.append(exel_list[i][0])
        for j in range(limit + 2):
            if exel_list[i][0] == exel_list[j + 2][0]:
                i += 1


    else:
        temp_list.append(exel_list[i][0])
