import openpyxl
from openpyxl.reader.excel import load_workbook
from rapidfuzz import fuzz


def read_programs_list():
    wb = load_workbook('backup/kru_full.xls')
    sheet = wb['Details']
    # sheet = wb.active
    # wb.get_sheet_by_name("Details")
    data = []
    for row in sheet.iter_rows(min_row=4, max_row=18713, min_col=1, max_col=7, values_only=True):
        data.append(row)
    return data


def read_program_list():
    wb = load_workbook('groups_test2.xlsx')
    sheet = wb['Sheet']
    # sheet = wb.active
    # wb.get_sheet_by_name("Details")
    data = []
    for row in sheet.iter_rows(min_row=1, max_row=23428, min_col=1, max_col=7, values_only=True):
        data.append(row)
    return data


def read_profile_programs():
    wb = load_workbook('profil1.xlsx')
    sheet = wb.active
    xlsx_seq = []
    for row in sheet.iter_rows(min_row=1, max_row=209, min_col=1, max_col=1, values_only=True):
        xlsx_seq.append(row)
    return xlsx_seq


def read_base_programs():
    wb = load_workbook('base.xlsx')
    sheet = wb.active
    xlsx_seq = []
    for row in sheet.iter_rows(min_row=1, max_row=185, min_col=1, max_col=1, values_only=True):
        xlsx_seq.append(row)
    return xlsx_seq


def write_to_exel_output(filtered_list, filename):
    workbook = openpyxl.Workbook()

    # Select the active worksheet
    worksheet = workbook.active

    for row_idx, row_data in enumerate(filtered_list, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

    # Save the workbook to an Excel file
    workbook.save(f'{filename}.xlsx')


# def check_item(sentence_, words_):
#     for j in range(len(words_)):
#         if sentence_.find(words_[j][0]) != -1:
#             return True


def check_miss_percent(sentence_, xlsx_sequence):
    for k in range(len(xlsx_sequence)):
        print(k)
        checker = fuzz.ratio(sentence_, xlsx_sequence[k][0])
        if checker > 50:
            print(xlsx_sequence[k][0])
            return xlsx_sequence[k][0]
        else:
            return False


def word_miss_percent(word, word_with):
    print(f"{word}=={word_with}")
    return fuzz.token_sort_ratio(word.encode('utf-8'), word_with.encode('utf-8'))
