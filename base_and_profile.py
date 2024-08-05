from openpyxl.reader.excel import load_workbook
from rapidfuzz import fuzz

from utils import read_profile_programs, read_base_programs, write_to_exel_output


def read_programs_list():
    wb = load_workbook('output_fil.xlsx')
    sheet = wb.active
    # sheet = wb.active
    # wb.get_sheet_by_name("Details")
    data = []
    for row in sheet.iter_rows(min_row=4, max_row=18713, min_col=1, max_col=8, values_only=True):
        data.append(row)
    return data


def check_equal_lists(value, list_of_values):
    for z in range(len(list_of_values)):
        if fuzz.token_sort_ratio(value.encode('utf-8'), list_of_values[z].encode('utf-8')) > 70:
            return True
        else:
            return False


exel_file = read_programs_list()
profile_list = read_profile_programs()
base_list = read_base_programs()
temp_list = []

print(base_list[0][0])
for i in range(len(exel_file)):
    if exel_file[i][0] is not None:
        for j in range(len(profile_list)):
            if fuzz.token_sort_ratio(exel_file[i][0].encode('utf-8'), profile_list[j][0].encode('utf-8')) > 88:
                temp_list.append((
                    exel_file[i][0], exel_file[i][1], exel_file[i][2], exel_file[i][3], exel_file[i][4],
                    exel_file[i][5], exel_file[i][6], "Профильное"))
                break
        for k in range(len(base_list)):
            if fuzz.token_sort_ratio(exel_file[i][0].encode('utf-8'), base_list[k][0].encode('utf-8')) > 88:
                temp_list.append((
                    exel_file[i][0], exel_file[i][1], exel_file[i][2], exel_file[i][3], exel_file[i][4],
                    exel_file[i][5], exel_file[i][6], "Базовое"))
                break

    else:
        temp_list.append(exel_file[i])

write_to_exel_output(temp_list, "base_and_profile")